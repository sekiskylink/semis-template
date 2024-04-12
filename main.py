import sys
import argparse
import dhis2.exceptions
import urllib3
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
from dhis2 import Api
from datetime import datetime, timedelta

urllib3.disable_warnings()
parser = argparse.ArgumentParser(
    prog="main.py",
    description="SEMIS Bulk Enrollment Template Generator",
    epilog="semis-template-generator",
)

parser.add_argument(
    "-P", "--program",
    help="The DHIS2 tracker program",
    dest="program",
    default="a6t4ASRXwPZ"
)

parser.add_argument(
    "-d", "--dhis2-url",
    help="The DHIS2 URL with out the trailing /api",
    dest="dhis2_url",
    default="https://emiseswatini.dev.hispuganda.org/emiseswatini"
)

parser.add_argument(
    "-u", "--username",
    help="DHIS2 username",
    dest="dhis2_user",
    default="admin"
)
parser.add_argument(
    "-p", "--password",
    help="The DHIS2 user's password",
    dest="dhis2_password",
    default="Dev2023!"
)

parser.add_argument(
    "-o", "--orgunit",
    help="The DHIS2 organisation unit ID. Comma separate multiple IDs",
    dest="orgunit",
    default=""
)
parser.add_argument(
    "-s", "--security-password",
    help="The password to protect the workbook and its sheets. Don't use the exclamation mark '!' Defaults to semis",
    dest="workbook_password",
    default="semis"
)

parser.add_argument(
    "-n", "--records",
    help="The Number of Students",
    dest="records",
    type= int,
    default=100
)
today = datetime.today()
two_years_ago = today - timedelta(days=730)
a_month_ahead = today + timedelta(days=30)
args = parser.parse_args()

program_fields: str = (
    "id,displayName,programType,"
    "programStages[id,displayName,programStageDataElements[compulsory,"
    "dataElement[id,name,formName,optionSetValue,optionSet[id,options[name]]]]],"
    "programTrackedEntityAttributes[trackedEntityAttribute"
    "[id,displayName,formName,generated,valueType,optionSet[id,options[name]],optionSetValue],mandatory]"
)
print("STARTING......")
api = Api(args.dhis2_url, args.dhis2_user, args.dhis2_password)

try:
    r = api.get(f'programs/{args.program}', params={'fields': program_fields})
except dhis2.exceptions.RequestException as e:
    print(e)
    sys.exit(1)

program = r.json()
# print(program["id"])
try:
    ou = api.get(f"organisationUnits", params={
        'fields': "id,name",
        'filter': f"id:in:[{args.orgunit}]",
    })
except dhis2.exceptions.RequestException as e:
    print(e)
    sys.exit(1)
org_units = ou.json()["organisationUnits"]
# print(org_units)
try:
    r = api.get(f"dataStore/semis/values", params={})
except dhis2.exceptions.RequestException as e:
    print(e)
    sys.exit(1)

dataStoreValues = r.json()

de_te_attribute_options = {}
column_optionSet = {}
compulsory_headers = ["orgUnitName", "orgUnit", "enrollmentDate"]


def column_number_to_name(column_number):
    column_name = ""
    while column_number > 0:
        column_number, remainder = divmod(column_number - 1, 26)
        column_name = chr(65 + remainder) + column_name
    return column_name


def get_stage_data_elements(stage: str, data_store_vals: list, program_stages: dict) -> list:
    if len(data_store_vals):
        program_stage_id = data_store_vals[0][stage]["programStage"]
        conf = list(filter(lambda x: x["id"] == program_stage_id, program_stages))
        ret = []
        if len(conf):
            des = conf[0]["programStageDataElements"]
            des.sort(key=lambda x: x["compulsory"], reverse=True)
            for de in des:
                column = f"{program_stage_id}.{de['dataElement']['id']}"
                # name = de['dataElement']['formName'] if 'formName' in de['dataElement'] else de['dataElement']['name']
                name = de['dataElement']['name'] if 'formName' in de['dataElement'] else de['dataElement']['name']
                ret.append((name, column))
                # track compulsory headers
                if de["compulsory"]:
                    compulsory_headers.append(column)
                # stealthily set the de, te options for later use
                if de['dataElement']['optionSetValue']: # we expect data to come from an option set
                    opts = [opt["name"] for opt in de["dataElement"]["optionSet"]["options"]]
                    opts.sort(key=lambda x: x)
                    de_te_attribute_options[column] = opts
                    # track option set for each column
                    column_optionSet[column] = de["dataElement"]["optionSet"]["id"]

        return ret
    return []


headers = ["#", "School Name", "School UID", "Enrollment Date"]
hidden_headers = ["ref", "orgUnitName", "orgUnit", "enrollmentDate"]
# append registration programStage.DE pairs for registration program
program_stages = program["programStages"]

headings = get_stage_data_elements("registration", dataStoreValues, program_stages)
for heading in headings:
    headers.append(heading[0])
    hidden_headers.append(heading[1])

# append TE attribute columns
tes = list(filter(
    lambda x: not (x["trackedEntityAttribute"]["generated"]) and
              x["trackedEntityAttribute"]["valueType"] not in ['IMAGE'],
    program["programTrackedEntityAttributes"])
)
tes.sort(reverse=True, key=lambda x: x["mandatory"])
# print("TES", tes)
if len(tes):
    for te in tes:
        headers.append(f"{te['trackedEntityAttribute']['displayName']}")
        te_id = f"{te['trackedEntityAttribute']['id']}"
        hidden_headers.append(te_id)
        if te['trackedEntityAttribute']['optionSetValue']:  # we expect data to come from an option set
            options = [opt["name"] for opt in te["trackedEntityAttribute"]["optionSet"]["options"]]
            options.sort(key=lambda x: x)
            de_te_attribute_options[te_id] = options
            column_optionSet[te_id] = te["trackedEntityAttribute"]["optionSet"]["id"]
            if te["mandatory"]:
                compulsory_headers.append(te_id)

headings = get_stage_data_elements("socio-economics", dataStoreValues, program_stages)
for heading in headings:
    headers.append(heading[0])
    hidden_headers.append(heading[1])

column_reference = {}
for idx, h in enumerate(hidden_headers, start=1):
    column_reference[h] = idx
# print("COLUMNS: ", column_reference)

wb = Workbook()
ws = wb.active
ws.title = "Data"
# wsprops = ws.sheet_properties
ws.row_dimensions.group(2, 2, hidden=True)
ws.column_dimensions.group("C","C", hidden=True)
ws.append(headers)
ws.append(hidden_headers)

for number in range(1, args.records + 1):
    ws[f'A{number + 2}'].value = f"{number}"
ou_names = [o["name"] for o in org_units]
# print(ou_names)

data_val = DataValidation(type="list", formula1='"' + ','.join(ou_names) + '"', allowBlank=True)
date_val = DataValidation(
    type="date", operator="between",
    formula1=f'DATE({two_years_ago.year},{two_years_ago.month},{two_years_ago.day})',
    formula2=f'DATE({a_month_ahead.year},{a_month_ahead.month},{a_month_ahead.day})', showDropDown=True)
date_val.prompt = f'Please enter a date between {two_years_ago.strftime("%Y-%m-%d")} and today!'
date_val.showInputMessage = True
date_val.showErrorMessage = True
date_val.error = f'The date must be between {two_years_ago.strftime("%Y-%m-%d")} and today!'
date_val.showError = True

ws.add_data_validation(data_val)
ws.add_data_validation(date_val)
offset = 2
data_val.add(f"B{offset + 1}:B{args.records + offset + 1}")
date_val.add(f"D{offset + 1}:D{args.records + offset + 1}")

ws1 = wb.create_sheet('MetaData')
ws1.sheet_state = 'hidden'
ws1.append(["programID","School Name", "School ID", "#"])
ws1.column_dimensions["A"].width = 15
ws1.column_dimensions["B"].width = 30
ws1.column_dimensions["C"].width = 15
ws1.column_dimensions["D"].width = 15

ws1.append([program["id"]])
for idx, ou in enumerate(org_units, start=2):
    ws1.cell(row=idx, column=2, value=ou["name"])
    ws1.cell(row=idx, column=3, value=ou["id"])

ws = wb["Data"]
for idx, number in enumerate(range(1, args.records + 1), start=3):
    ws.cell(row=idx, column=3).value = f'=IFERROR(VLOOKUP(B{idx},MetaData!$B$2:$C${len(org_units) + 1},2,FALSE), "")'


def contains_comma(string_list):
    return any(',' in s for s in string_list)


long_option_sets = {}
for k, v in de_te_attribute_options.items():
    column_label = column_number_to_name(column_reference[k])
    if len(','.join(v)) <= 256 and not contains_comma(v): # limit for dynamic list
        data_val = DataValidation(type="list", formula1='"' + ','.join(v) + '"', allowBlank=True)
        ws.add_data_validation(data_val)
        offset = 2
        data_val.add(f"{column_label}{offset + 1}:{column_label}{args.records + offset}")
    else:
        # print(k, "=>", v, ">>", column_label, "===>", len(v), f"optionSet={column_optionSet[k]}")
        option_set = column_optionSet[k]
        if option_set not in long_option_sets:
            long_option_sets[option_set] = {
                "options": v,
                "appliesTo": [column_label]
            }
        else:
            long_option_sets[option_set]["appliesTo"].append(column_label)

# keep these optionset options in metadata sheet
for idx, (option_set, val) in enumerate(long_option_sets.items(), start=1):
    ws1 = wb['MetaData']
    metadata_sheet_column_offset = 4
    ws1.cell(row=1, column=metadata_sheet_column_offset + idx, value=f"optionSet.{option_set}")
    max_option_length = 8
    for i, option in enumerate(val["options"], start=2):
        ws1.cell(row=i, column=metadata_sheet_column_offset + idx, value=option)
        max_option_length = max(max_option_length, len(option))

    options_col = column_number_to_name(metadata_sheet_column_offset + idx)
    ws1.column_dimensions[options_col].width = max_option_length

    option_set_column_label = column_number_to_name(metadata_sheet_column_offset + idx)
    dv = DataValidation(
        type="list",
        formula1=f'=MetaData!${option_set_column_label}$2:${option_set_column_label}${len(val["options"])}',
        allowBlank=True
    )
    ws = wb['Data']
    ws.add_data_validation(dv)
    for col in val["appliesTo"]:
        # add validation to each of the columns it applies
        dv.add(f"{col}3:{col}{args.records + 2}")

# style the headers
ws = wb['Data']
ft = Font(bold=True, size=14)
for col, h in enumerate(headers, start=1):
    column_label = column_number_to_name(col)
    cell = ws[f"{column_label}1"]
    cell.font = ft
    if len(h) < 8:
        ws.column_dimensions[column_label].width = 11
    else:
        ws.column_dimensions[column_label].width = len(h) + 3
ws.column_dimensions["B"].width = 40

ft = Font(bold=True, size=14, color="FF0000")
# print("COMPULSORY HEADERS ", compulsory_headers)
for h in compulsory_headers:
    try:
        col_idx = hidden_headers.index(h) + 1
        # print(f"HEADER_INDEX: {h}", hidden_headers.index(h))
        cell = ws[f"{column_number_to_name(col_idx)}1"]
        cell.font = ft
    except ValueError:
        pass

wb.security.workbookPassword = 'semis'
wb.security.lockStructure = True
ws1 = wb["MetaData"]
ws1.protection.sheet = True
ws1.protection.enable()
ws1.protection.password = "semis"
wb.save('SemisTemplate.xlsx')
print("DONE..........")